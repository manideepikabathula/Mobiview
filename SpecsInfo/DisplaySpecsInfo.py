import re
import logging
from Lib.ParserUtils import ParserUtils
from SpecsInfo.MobileSpecsInfo import MobileSpecsInfo
from openpyxl.utils import get_column_letter
from Lib.FileSystemUtils import FileSystemUtils

log = logging.getLogger(__name__)


class DisplaySpecsInfo(MobileSpecsInfo):
    def __init__(self):
        super().__init__()
        self.Display_Density = None
        self.Screen_Size = None
        self.Screen_Brightness = None
        self.Refresh_Rate = None
        self.Screen_Off_Timeout = None
        self.Screen_Rotation = None

    def grepInfo(self):
        self.getDisplayDensity()
        self.getScreenSize()
        self.getScreenBrightness()
        self.getRefreshRate()
        self.getScreenOffTimeout()
        self.getScreenRotation()
        return self.DisplaySpecsInfoDict

    def cleanup(self):
        pass

    def getDisplayDensity(self):
        self.command = self.ADBObj.getADBWindowsManagerCommand() + ' density '
        self.DisplayDensity = self.executeCommandOnDevice(command=self.command)
        pattern = re.compile(r':\s+(?P<Display_Density>.*)')
        rvalue = ParserUtils.parseDataViaRegex(pattern, self.DisplayDensity)
        self.updateDictionary(dictName=self.DisplaySpecsInfoDict, key='DisplayDensity',
                              value=rvalue.get('Display_Density'))
        return rvalue.get('Display_Density')

    def getScreenSize(self):
        self.command = self.ADBObj.getADBWindowsManagerCommand() + ' size '
        self.ScreenSize = self.executeCommandOnDevice(command=self.command)
        pattern = re.compile(r':\s+(?P<screen_size>.*)')
        rvalue = ParserUtils.parseDataViaRegex(pattern, self.ScreenSize)
        self.updateDictionary(dictName=self.DisplaySpecsInfoDict, key='DisplayScreenSize',
                              value=rvalue.get('screen_size'))
        return rvalue.get('screen_size')

    def getScreenBrightness(self):
        self.command = self.ADBObj.getADBShellCommand() + ' settings get system screen_brightness'
        self.ScreenBrightness = self.executeCommandOnDevice(command=self.command).strip()
        self.updateDictionary(dictName=self.DisplaySpecsInfoDict, key='ScreenBrightness',
                              value=self.ScreenBrightness)
        return self.ScreenBrightness

    def getRefreshRate(self):
        # Execute the command to get the refresh rate
        self.command = self.ADBObj.getADBShellCommand() + ('dumpsys display | grep -E "mDefaultPeakRefreshRate" | head '
                                                           '-n 1')
        self.RefreshRate = self.executeCommandOnDevice(command=self.command).strip()

        # Extract the refresh rate using regex
        pattern = re.compile(r'mDefaultPeakRefreshRate= ?(?P<refresh_rate>[\d\.]+)')
        rvalue = ParserUtils.parseDataViaRegex(pattern, self.RefreshRate)
        refresh_rate = rvalue.get('refresh_rate') if rvalue else None

        # Update the dictionary with the refresh rate
        self.updateDictionary(dictName=self.DisplaySpecsInfoDict, key='RefreshRate',
                              value=refresh_rate)
        return refresh_rate

    def getScreenOffTimeout(self):
        self.command = self.ADBObj.getADBShellCommand() + ' settings get system screen_off_timeout'
        self.ScreenOffTimeout = self.executeCommandOnDevice(command=self.command).strip()
        self.updateDictionary(dictName=self.DisplaySpecsInfoDict, key='ScreenOffTimeout',
                              value=self.ScreenOffTimeout)
        return self.ScreenOffTimeout

    def getScreenRotation(self):
        self.command = self.ADBObj.getADBShellCommand() + ' settings get system user_rotation'
        self.ScreenRotation = self.executeCommandOnDevice(command=self.command).strip()
        self.updateDictionary(dictName=self.DisplaySpecsInfoDict, key='ScreenRotation',
                              value=self.ScreenRotation)
        return self.ScreenRotation

    def generateXLSXReport(self, xlsObj=None, wb=None, ws=None, dataDict=None):
        headers = ["Parameters", "Description", "Results"]

        for idx, header in enumerate(headers):
            cellref = ws.cell(row=2, column=idx + 2)
            ws.column_dimensions[get_column_letter(idx + 2)].width = 40
            cellref.style = xlsObj.getNamedStyle(stylename="headerRow")
            cellref.value = header

        param_descriptions = {
            "DisplayDensity": "The density of the display, measured in DPI",
            "DisplayScreenSize": "The size of the screen, typically measured diagonally in inches",
            "ScreenBrightness": "The current screen brightness level, measured on a scale from 1 to 255",
            "RefreshRate": " This is the maximum rate at which the screen can refresh its content, measured in Hertz (Hz)",
            "ScreenOffTimeout": "The duration before the screen turns off automatically, measured in milliseconds",
            "ScreenRotation": "The current orientation of the screen, represented as degrees (0, 90, 180, 270)",
            # Add a default description for missing keys
            "<UNKNOWN_KEY>": "Description not available"
        }
        Screen_Rotation_descriptions = {
            "0": "The screen is in its default portrait orientation.",
            "90": "The screen is rotated 90 degrees clockwise, in landscape mode.",
            "180": "The screen is upside down, in reverse portrait mode.",
            "270": "The screen is rotated 90 degrees counterclockwise, in reverse landscape mode."
        }

        for row_idx, (key, value) in enumerate(dataDict.items(), start=3):
            # Parameters
            param_cell = ws.cell(row=row_idx, column=2)
            param_cell.style = xlsObj.getNamedStyle(stylename="normalRow")
            param_cell.value = key

            # Description
            desc_cell = ws.cell(row=row_idx, column=3)
            desc_cell.style = xlsObj.getNamedStyle(stylename="normalRow")
            if key == "ScreenRotation":
                desc_cell.value = Screen_Rotation_descriptions.get(value, "Description not available")
            else:
                desc_cell.value = param_descriptions.get(key, "Description not available")

            # Results
            result_cell = ws.cell(row=row_idx, column=4)
            result_cell.style = xlsObj.getNamedStyle(stylename="normalRow")
            charlist = ["[", "'", "]"]
            result_value = FileSystemUtils.replaceChars(value, charlist)
            result_cell.value = str(result_value)

            # Debugging prints
            print(f"Processing key: {key}")
            print(f"Value: {result_value}")
            print(f"Description: {desc_cell.value}")

        col_idx = 2
        last_row_idx = len(dataDict) + 3
        for ctr in range(col_idx, col_idx + 3):
            cellref = ws.cell(row=last_row_idx, column=ctr)
            cellref.style = xlsObj.getNamedStyle(stylename="lastRow")
